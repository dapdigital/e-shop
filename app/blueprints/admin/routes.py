import os
from werkzeug.utils import secure_filename
from flask import render_template, redirect, url_for, flash, request, current_app
from . import admin_bp
from flask_login import login_required
from .decorators import admin_requerido
from app import db
from app.models import Categoria, Usuario, Pedido, Producto, MensajeContacto
from .forms import FormCategoria, FormProducto


@admin_bp.route('/dashboard')
@login_required
@admin_requerido
def dashboard():
    total_ventas = db.session.query(db.func.sum(Pedido.total)).filter(
        Pedido.estado.in_(['pagado', 'enviado', 'entregado'])
    ).scalar() or 0

    total_pedidos = Pedido.query.count()
    pedidos_pendientes = Pedido.query.filter_by(estado='pendiente').count()

    productos_bajo_stock = Producto.query.filter(
        Producto.activo == True,
        Producto.stock <= 5
    ).order_by(Producto.stock.asc()).all()

    mensajes_sin_leer = MensajeContacto.query.filter_by(leido=False).count()

    return render_template('admin/home.html',
        total_ventas=total_ventas,
        total_pedidos=total_pedidos,
        pedidos_pendientes=pedidos_pendientes,
        productos_bajo_stock=productos_bajo_stock,
        mensajes_sin_leer=mensajes_sin_leer
    )


# ==================== CATEGORÍAS ====================

@admin_bp.route('/categorias')
@login_required
@admin_requerido
def listar_categorias():
    categorias = Categoria.query.order_by(Categoria.nombre).all()
    return render_template('admin/categorias/listar.html', categorias=categorias)


@admin_bp.route('/categorias/crear', methods=['GET', 'POST'])
@login_required
@admin_requerido
def crear_categoria():
    form = FormCategoria()

    if form.validate_on_submit():
        nueva = Categoria(
            nombre=form.nombre.data,
            descripcion=form.descripcion.data,
            activa=True
        )
        db.session.add(nueva)
        db.session.commit()
        flash('Categoría creada correctamente.', 'success')
        return redirect(url_for('admin.listar_categorias'))

    return render_template('admin/categorias/formulario.html', form=form, titulo='Nueva categoría')


@admin_bp.route('/categorias/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_requerido
def editar_categoria(id):
    categoria = Categoria.query.get_or_404(id)
    form = FormCategoria(obj=categoria)

    if form.validate_on_submit():
        categoria.nombre = form.nombre.data
        categoria.descripcion = form.descripcion.data
        categoria.activa = form.activa.data
        db.session.commit()
        flash('Categoría actualizada correctamente.', 'success')
        return redirect(url_for('admin.listar_categorias'))

    return render_template('admin/categorias/formulario.html', form=form, titulo='Editar categoría')


@admin_bp.route('/categorias/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_requerido
def eliminar_categoria(id):
    categoria = Categoria.query.get_or_404(id)
    categoria.activa = False
    db.session.commit()
    flash('Categoría desactivada.', 'warning')
    return redirect(url_for('admin.listar_categorias'))


# ==================== CLIENTES ====================

@admin_bp.route('/gestion-clientes')
@login_required
@admin_requerido
def gestion_clientes():
    clientes = Usuario.query.filter_by(rol='cliente').order_by(Usuario.nombre).all()
    return render_template('admin/clientes/listar.html', clientes=clientes)


@admin_bp.route('/gestion-clientes/toggle/<int:id>', methods=['POST'])
@login_required
@admin_requerido
def toggle_cliente(id):
    cliente = Usuario.query.get_or_404(id)
    cliente.activo = not cliente.activo
    db.session.commit()
    estado = 'activado' if cliente.activo else 'desactivado'
    flash(f'Cliente {estado} correctamente.', 'success')
    return redirect(url_for('admin.gestion_clientes'))


# ==================== PEDIDOS ====================

@admin_bp.route('/gestion-pedidos')
@login_required
@admin_requerido
def gestion_pedidos():
    pedidos = Pedido.query.order_by(Pedido.fecha.desc()).all()
    return render_template('admin/pedidos/listar.html', pedidos=pedidos)


@admin_bp.route('/gestion-pedidos/estado/<int:id>', methods=['POST'])
@login_required
@admin_requerido
def cambiar_estado_pedido(id):
    pedido = Pedido.query.get_or_404(id)
    orden_estados = ['pendiente', 'pagado', 'enviado', 'entregado']

    if pedido.estado in orden_estados:
        idx = orden_estados.index(pedido.estado)
        if idx < len(orden_estados) - 1:
            pedido.estado = orden_estados[idx + 1]
            db.session.commit()
            flash(f'Pedido actualizado a "{pedido.estado}".', 'success')
        else:
            flash('El pedido ya está en el último estado.', 'info')

    return redirect(url_for('admin.gestion_pedidos'))


# ==================== PRODUCTOS ====================

@admin_bp.route('/productos')
@login_required
@admin_requerido
def productos():
    lista = Producto.query.order_by(Producto.nombre).all()
    return render_template('admin/productos/listar.html', productos=lista)


@admin_bp.route('/productos/crear', methods=['GET', 'POST'])
@login_required
@admin_requerido
def crear_producto():
    form = FormProducto()
    form.categoria_id.choices = [(c.id, c.nombre) for c in Categoria.query.filter_by(activa=True).all()]

    if form.validate_on_submit():
        nombre_archivo = None
        archivo = request.files.get('imagen')

        if archivo and archivo.filename:
            nombre_archivo = secure_filename(archivo.filename)
            ruta_guardado = os.path.join(current_app.root_path, 'static', 'img', nombre_archivo)
            archivo.save(ruta_guardado)

        nuevo = Producto(
            nombre=form.nombre.data,
            descripcion=form.descripcion.data,
            precio=form.precio.data,
            stock=form.stock.data,
            categoria_id=form.categoria_id.data,
            imagen=nombre_archivo,
            activo=True
        )
        db.session.add(nuevo)
        db.session.commit()
        flash('Producto creado correctamente.', 'success')
        return redirect(url_for('admin.productos'))

    return render_template('admin/productos/formulario.html', form=form, titulo='Nuevo producto')


@admin_bp.route('/productos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_requerido
def editar_producto(id):
    producto = Producto.query.get_or_404(id)
    form = FormProducto(obj=producto)
    form.categoria_id.choices = [(c.id, c.nombre) for c in Categoria.query.filter_by(activa=True).all()]

    if form.validate_on_submit():
        archivo = request.files.get('imagen')

        if archivo and archivo.filename:
            nombre_archivo = secure_filename(archivo.filename)
            ruta_guardado = os.path.join(current_app.root_path, 'static', 'img', nombre_archivo)
            archivo.save(ruta_guardado)
            producto.imagen = nombre_archivo

        producto.nombre = form.nombre.data
        producto.descripcion = form.descripcion.data
        producto.precio = form.precio.data
        producto.stock = form.stock.data
        producto.categoria_id = form.categoria_id.data
        producto.activo = form.activo.data
        db.session.commit()
        flash('Producto actualizado correctamente.', 'success')
        return redirect(url_for('admin.productos'))

    return render_template('admin/productos/formulario.html', form=form, titulo='Editar producto')


@admin_bp.route('/productos/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_requerido
def eliminar_producto(id):
    producto = Producto.query.get_or_404(id)
    producto.activo = False
    db.session.commit()
    flash('Producto desactivado.', 'warning')
    return redirect(url_for('admin.productos'))

@admin_bp.route('/productos/importar', methods=['GET', 'POST'])
@login_required
@admin_requerido
def importar_productos():
    if request.method == 'POST':
        archivo = request.files.get('archivo_csv')

        if not archivo or not archivo.filename:
            flash('Selecciona un archivo CSV o Excel.', 'danger')
            return redirect(url_for('admin.importar_productos'))

        nombre_archivo = archivo.filename.lower()
        filas = []

        if nombre_archivo.endswith('.xlsx'):
            from openpyxl import load_workbook

            try:
                wb = load_workbook(archivo, data_only=True)
                hoja = wb.active
                encabezados = [str(c.value).strip().lower() if c.value else '' for c in hoja[1]]

                for fila_excel in hoja.iter_rows(min_row=2, values_only=True):
                    fila_dict = dict(zip(encabezados, fila_excel))
                    filas.append(fila_dict)
            except Exception:
                flash('No se pudo leer el archivo Excel. Verifica que no esté dañado.', 'danger')
                return redirect(url_for('admin.importar_productos'))

        elif nombre_archivo.endswith('.csv'):
            import csv, io

            try:
                contenido = archivo.stream.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                flash('No se pudo leer el archivo. Guárdalo como CSV UTF-8 desde Excel.', 'danger')
                return redirect(url_for('admin.importar_productos'))

            lector = csv.DictReader(io.StringIO(contenido))
            filas = list(lector)

        else:
            flash('El archivo debe ser .csv o .xlsx', 'danger')
            return redirect(url_for('admin.importar_productos'))

        columnas_esperadas = {'categoria', 'nombre', 'descripcion', 'precio', 'stock', 'imagen'}
        if not filas or not columnas_esperadas.issubset(set(filas[0].keys())):
            flash(f'El archivo debe tener las columnas: {", ".join(columnas_esperadas)}', 'danger')
            return redirect(url_for('admin.importar_productos'))

        creados = 0
        actualizados = 0
        categorias_nuevas = 0
        errores = []

        for i, fila in enumerate(filas, start=2):
            nombre_cat = str(fila.get('categoria') or '').strip()
            nombre_prod = str(fila.get('nombre') or '').strip()

            if not nombre_cat or not nombre_prod:
                errores.append(f'Fila {i}: falta categoría o nombre, se omitió.')
                continue

            try:
                precio = float(fila.get('precio') or 0)
                stock = int(float(fila.get('stock') or 0))
            except (ValueError, TypeError):
                errores.append(f'Fila {i}: precio o stock inválido, se omitió.')
                continue

            categoria = Categoria.query.filter(
                db.func.lower(Categoria.nombre) == nombre_cat.lower()
            ).first()

            if not categoria:
                categoria = Categoria(nombre=nombre_cat, activa=True)
                db.session.add(categoria)
                db.session.flush()
                categorias_nuevas += 1

            producto = Producto.query.filter(
                db.func.lower(Producto.nombre) == nombre_prod.lower(),
                Producto.categoria_id == categoria.id
            ).first()

            imagen = str(fila.get('imagen') or '').strip() or None
            descripcion = str(fila.get('descripcion') or '').strip()

            if producto:
                producto.descripcion = descripcion
                producto.precio = precio
                producto.stock = stock
                if imagen:
                    producto.imagen = imagen
                actualizados += 1
            else:
                nuevo = Producto(
                    nombre=nombre_prod,
                    descripcion=descripcion,
                    precio=precio,
                    stock=stock,
                    imagen=imagen,
                    categoria_id=categoria.id,
                    activo=True
                )
                db.session.add(nuevo)
                creados += 1

        db.session.commit()

        resumen = f'Importación completa: {creados} productos creados, {actualizados} actualizados, {categorias_nuevas} categorías nuevas.'
        flash(resumen, 'success')
        for err in errores:
            flash(err, 'warning')

        return redirect(url_for('admin.productos'))

    return render_template('admin/productos/importar.html')


# ==================== MENSAJES DE CONTACTO ====================

@admin_bp.route('/mensajes')
@login_required
@admin_requerido
def gestion_mensajes():
    mensajes = MensajeContacto.query.order_by(MensajeContacto.fecha.desc()).all()
    return render_template('admin/mensajes/listar.html', mensajes=mensajes)


@admin_bp.route('/mensajes/leido/<int:id>', methods=['POST'])
@login_required
@admin_requerido
def marcar_mensaje_leido(id):
    mensaje = MensajeContacto.query.get_or_404(id)
    mensaje.leido = not mensaje.leido
    db.session.commit()
    return redirect(url_for('admin.gestion_mensajes'))


@admin_bp.route('/mensajes/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_requerido
def eliminar_mensaje(id):
    mensaje = MensajeContacto.query.get_or_404(id)
    db.session.delete(mensaje)
    db.session.commit()
    flash('Mensaje eliminado.', 'info')
    return redirect(url_for('admin.gestion_mensajes'))